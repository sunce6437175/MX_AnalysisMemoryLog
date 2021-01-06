#!/usr/bin/python
# -*- coding: UTF-8 -*-

# WDXStability report import
from __future__ import print_function
import os, mimetypes
import datetime
import json
import time
import xlsxwriter
import xlrd
import openpyxl
import xlwings as xw
from collections import deque
# Send txt&excel import
import smtplib,time
import datetime
import schedule
from email.mime.text import MIMEText  
from email.mime.image import MIMEImage
from email.mime.multipart import MIMEMultipart
from email.header import Header
from email.mime.base import MIMEBase
from email.mime.application import MIMEApplication
from PIL import ImageGrab

# 获取中国节假日api
from chinese_calendar import is_workday

# 配置参数路径class
class KeyType:
    # 空间占用关键字
    spaceUsage = "total"
    # 搜索内存关键字
    deWeightTime = "top -m | grep MXNavi"
    # #有效内容关键字
    keyMXNavi = "/usr/bin/MXNavi"
    # 最新批量json配置项 (从本地改为服务器路径)
    configJsonName = "//192.168.2.22/cns3.0_sop2_ma/04.C Sample/03.非功能测试/稳定性测试/MX_AnalysisMemoryLog/config/config.json"
    # 最新批量json配置项地址获取(获取当前文件的绝对路径)
    configJsonPath = os.path.join(os.path.abspath(os.path.dirname(__file__)),configJsonName).replace("\\",'/')
    # 修改添加自动拾取时间的json配置项后写入文件 (从本地改为服务器路径)
    testconfigJsonName = "//192.168.2.22/cns3.0_sop2_ma/04.C Sample/03.非功能测试/稳定性测试/MX_AnalysisMemoryLog/config/testJson.json"
    # 修改添加自动拾取时间的json配置项后写入文件路径
    testconfigJsonPath = os.path.join(os.path.abspath(os.path.dirname(__file__)),testconfigJsonName).replace("\\",'/')
    # 批量判断文件路径
    timelineName = "//192.168.2.22/cns3.0_sop2_ma/04.C Sample/03.非功能测试/稳定性测试/"
    
    timelinePath = os.path.join(os.path.abspath(os.path.dirname(__file__)),timelineName).replace("\\",'/')

class Analysism:
    def __init__(self,name,startTime,finishTime,readPath,writePath,exclPath,KPI,secondaryMaximum,divide_The_Value,divide_The_Time,pState,SpaceUsageKPI):
        self.name = name
        self.startTime = startTime
        self.finishTime = finishTime
        self.readPath = readPath
        self.writePath = writePath
        self.exclPath = exclPath
        self.KPI = KPI
        self.secondaryMaximum = secondaryMaximum
        self.divide_The_Value = divide_The_Value
        self.divide_The_Time = divide_The_Time
        self.pState = pState
        self.SpaceUsageKPI = SpaceUsageKPI

    startLineNum = 0    # 开始行数
    finishLineNum = 0   # 结束行数

    sheetcount = 0
    sheetNameList = []          # 保汇总sheet姓名
    starNumlist = []
    endNumlist = []
    maxNumlist = []
    test_Result_status = []
    test_Result = ""

    effective_running_time = []     #导航有效运行时间
    timestamp = []                 #超1G时间戳
    error_running_time = []       #（超1G达到时间 和 第一次超1G的时间戳 + 当在1000-1024之间长时间保持 + 未超1G但开始结束的落差值在300MB以上）  
    divide_Time_list = []         #超过在1000 - 1024之间的连续时间耗时
    surpassTimeS = 0              #超过在1024放置时间
    spaceUsageList = []         #空间占用
    spaceUsageNum = []

    def check_tmpLine(self):
        tempLienNum = 0     # 标注行数
        if os.path.exists(self.readPath):
            with open(self.readPath,'r',encoding = "UTF-8",errors = "ignore") as read_file:
                for line in read_file:
                    tempLienNum = tempLienNum + 1
                    if self.startTime in line:
                        if KeyType.deWeightTime in line:
                            continue
                        else:
                            Analysism.startLineNum = tempLienNum
                            
                    if self.finishTime in line:
                        if "END" in line:
                            continue
                        else:
                            Analysism.finishLineNum = tempLienNum
        else:
            print('%s 此log不存在'%(self.readPath))
            # continue

        return(Analysism.startLineNum,Analysism.finishLineNum)

    #判断有效运行时间
    def check_memoryTime(self):
        tempList = []       # 标注列表
        tempLienNum = 0     # 标注行数
        tempNum = 0         # 对比行数
    
        vSt = datetime.datetime.strptime(self.startTime.replace("/",'-'),"%Y-%m-%d %H:%M:%S")
        vFt = datetime.datetime.strptime(self.finishTime.replace("/",'-'),"%Y-%m-%d %H:%M:%S")
        # print("开始时间：%s 和对应格式 %s"%(vSt,type(vSt)))
        # print("结束时间：%s 和对应格式 %s"%(vFt,type(vFt)))
        # 2.1）抽取开始和结束的时间戳，判断有效运行时间（单位：小时）
        dayCtimes = ((vFt - vSt).total_seconds())/3600      # dayCtimes = ((vSt - vFt).total_seconds())/3600
        # print("本次有效时间-----共%.2f小时-----"%(dayCtimes))
        Analysism.effective_running_time.append("运行 %.2f H"%(dayCtimes))
        # 2.2）以及当内存超1024MB时所需时间。（超 kpi xxxxMB才需判断）
        stmpLine = self.startLineNum      # 调用父类check_tmpLine方法并赋值起始行，防止初始化循环调用
        ftmpLine = self.finishLineNum      # 调用父类check_tmpLine方法并赋值结束行，防止初始化循环调用
        
        if stmpLine <= ftmpLine and ftmpLine > stmpLine:
            with open(self.readPath,'r',encoding='UTF-8',errors="ignore") as read_file:
                for xline in read_file:
                    xline = xline.strip('\n')
                    if tempNum >= stmpLine and tempNum < ftmpLine:
                        if KeyType.deWeightTime in xline:
                            continue
                        else:
                            a = xline.split()
                            
                            if len(a) > 5 :
                                if KeyType.keyMXNavi in a:
                                    b = a[5]
                                    if int(b[:-1]) >= int(self.KPI):
                                        surpassDay = str((a[0]).strip('['))
                                        surpassHour = (a[1])
                                        surpassStr = (surpassDay + ' '+ surpassHour).rsplit(']')
                                        surpassTime = datetime.datetime.strptime((surpassStr[0]).replace("/",'-'),"%Y-%m-%d %H:%M:%S")

                                        Analysism.timestamp.append(a[0] + ' ' + a[1])
                                        Analysism.surpassTimeS = ((surpassTime - vSt).seconds)/3600

                                        # print("导航放置 %.2f 小时到达 %s MB"%(surpassTimeS,self.kPI))
                                        Analysism.error_running_time.append("超过%dMB的时间戳为%s/导航放置%.2f小时到达%dMB"%(self.KPI,a[0] + ' ' +a[1],Analysism.surpassTimeS,self.KPI))
                                        break
                                    else:
                                        pass
                                        # print("不知道啥数据%s"%(a))
                                
                                else:
                                    pass
                                    # print("无效数据%s"%(a))
                            elif KeyType.spaceUsage in a:
                                c = a[2]
                                Analysism.spaceUsageList.append(int(c)/1024/1024)
                                # print("空间占用数据：%s"%(a))
                                # print("空间占用个数：%s"%len(a))
                                # print("该行数的超过边界小于5的所在行%s"%(a))


                    else:
                        pass
                    tempNum = tempNum + 1
        else:
            pass

        return(Analysism.effective_running_time,Analysism.timestamp,Analysism.error_running_time,Analysism.spaceUsageList)
    # 判断空间占用的开始和结束及最大和平均值
    def check_spaceUsage(self):
        tempLienNum = 0        # 标注行数
        tempList = []          # 标注列表
        tempNum = 0            # 对比行数
        KPIList = []         # 获取超KPI的数据列表
        stempList = []
        stmpLine = Analysism.spaceUsageList      # 调用父类check_tmpLine方法并赋值起始行，防止初始化循环调用

        tempList = stmpLine
        startNum = float('%.2f' %(tempList[0]))
        endNum = float('%.2f' %(tempList[-1]))
        maxNum = float('%.2f' %(max(tempList))) 

        startNumstr = '开始占用空间' + str(tempList[0]) + ' GB'
        endNumstr = '结束占用空间' + str(tempList[-1]) + ' GB'
        maxNumastr = '最大占用空间' + str(max(tempList)) + ' GB'

        Analysism.spaceUsageNum.append(max(tempList))
        if maxNum >= float(self.SpaceUsageKPI) :
            # print("空间占用NG %s"%(maxNum))
            with open(self.readPath,'r',encoding='UTF-8',errors="ignore") as read_file:
                for xline in read_file:
                    xline = xline.strip('\n')

                    if KeyType.deWeightTime in xline:
                        continue
                    else:
                        a = xline.split()
                        if len(a) < 5 :
                            if KeyType.spaceUsage in a:
                                b = float(a[2])/1024/1024
                                
                                if b >= float(self.SpaceUsageKPI):
                                    print("超%s的所行%s 和大小值%.2f"%(self.SpaceUsageKPI,int(tempLienNum),b))
                                else:
                                    continue
                    tempLienNum = tempLienNum + 1


        return (Analysism.spaceUsageNum)


    # 判断内存开始和结束以及最大值
    def check_memorylog(self):
        tempLienNum = 0        # 标注行数
        tempList = []          # 标注列表
        tempNum = 0            # 对比行数
        KPIList = []         # 获取超KPI的数据列表

        stempLienNum = 0
        stempList = []
        stempNum = 0
        stempNumOneT = 0
        stempNumi = 0
        
        dtvList = []

        #取得开始结束值以及最大值 
        stmpLine = Analysism.startLineNum      # 调用父类check_tmpLine方法并赋值起始行，防止初始化循环调用
        ftmpLine = Analysism.finishLineNum      # 调用父类check_tmpLine方法并赋值结束行，防止初始化循环调用

        if stmpLine <= ftmpLine and ftmpLine > stmpLine:
            with open(self.readPath,'r',encoding='UTF-8',errors="ignore") as read_file:
                for xline in read_file:
                    xline = xline.strip('\n')
                    if tempNum >= stmpLine and tempNum < ftmpLine:
                        if KeyType.deWeightTime in xline:
                            continue
                        else:
                            a = xline.split()
                            if len(a) > 5 :
                                if KeyType.keyMXNavi in a:
                                    b = a[5]
                                    tempList.append(int(b[:-1]))
                                else:
                                    pass
                            else:
                                pass
                    else:
                        pass
                    tempNum = tempNum + 1
        else:
            print("！输入参时间错误！")
        print("起始内存值：%s MB,结束内存值：%s MB,最大内存值：%s MB"%(tempList[0],tempList[-1],max(tempList)))
        sNum = tempList[0]
        eNum = tempList[-1]
        mNum  = max(tempList)

        Analysism.starNumlist.append(sNum)
        Analysism.endNumlist.append(eNum)
        Analysism.maxNumlist.append(mNum)
        startNum = '起始内存值: ' + str(tempList[0]) + ' MB'
        endNum = '结束内存值: ' + str(tempList[-1]) + ' MB'
        maxNum = '最大内存值: ' + str(max(tempList)) + ' MB'

        # 实现场景1：判断峰值或结束值是已超1024MB
        if mNum >= int(self.KPI) or eNum >= int(self.KPI):
            print("实现场景1：判断峰值和结束已超1024MB")
            self.write_to_time()
            Analysism.sheetNameList.append(self.name)
            # 不准确待验证,会循环添加有问题
            # Analysism.error_running_time.append("超过%dMB的时间戳为%s/导航放置%.2f小时到达%dMB"%(self.KPI,Analysism.timestamp,Analysism.surpassTimeS,self.KPI))
            if self.pState == 'normal':
                Analysism.test_Result = 'NG'
                Analysism.test_Result_status.append(Analysism.test_Result) 
            # 取出全部数据生成Excel sheet    
            if stmpLine <= ftmpLine and ftmpLine > stmpLine:
                with open(self.readPath,'r',encoding='UTF-8',errors="ignore") as read_file:
                    for kline in read_file:
                        if "BEGIN" in kline:
                            continue
                        elif "END" in kline:
                            continue
                        else:
                            kline = kline.strip('\n').split()
                            KPIList.append(kline)
            else:
                print("起始结束行位置错误")

        # 实现场景2：未超1024MB，但长时间保持在1000MB，也就是在1000-1024之间长时间保持（保持时间 暂定≥60s，后期改为可配置）
        elif int(self.secondaryMaximum) <= mNum < int(self.KPI) or int(self.secondaryMaximum) <= eNum < int(self.KPI) :
            if stmpLine <= ftmpLine and ftmpLine > stmpLine:
                with open(self.readPath,'r',encoding='UTF-8',errors="ignore") as read_file:
                    i = 0
                    b = 0
                    for sline in read_file:
                        sline = sline.strip('\n')
                        if stempNum >= stmpLine and stempNum < ftmpLine:
                            sa = sline.split()
                            sb = sa[5]
                            stempNumOneT = int(sb[:-1])
                            # 判断范围在1000 - 1024之间的时间和信息   方法：不连续，用“1”来间隔
                            if stempNumOneT >= int(self.secondaryMaximum) and stempNumOneT < int(self.KPI):
                                # 加入所在目标行位置
                                sa.append(i+1)
                                stempList.append(sa)
                                # print("1000<=X<1024 所在行位置： %s  的行数 = %d"%(sa,i+1))
                                b = b+1
                            else:
                                pass
                                # print("不在1000<=X<1024范围内")
                        else:
                            print("！输入参数错误！")
                        stempNum = stempNum + 1
                        i = i +1
                        if i != ftmpLine :
                            continue
                        elif i == ftmpLine + 1 :
                            break
                        print('超过在1000 - 1024之间的连续时间：%ds'%b)

                        if b >= int(self.divide_The_Time) :
                            # Analysism.divide_Time_list.append("超过在1000 - 1024之间的连续时间：%ds'"%b)
                            Analysism.error_running_time.append("超过在1000 - 1024之间的连续时间%ds'"%b)
                            print("实现场景2：未超1024MB，但长时间保持在1000MB，也就是在1000-1024之间长时间保持（保持时间 暂定≥%ds）"%int(self.divide_The_Time))
                            self.write_to_time()
                            Analysism.sheetNameList.append(self.name)
                            
                            if self.pState == 'normal':
                                Analysism.test_Result = 'NG'
                                Analysism.test_Result_status.append(Analysism.test_Result) 
                        else:
                            print("未实现场景2：未超1024MB，但长时间保持在1000MB，也就是在1000-1024之间长时间保持未≥%ss "%self.divide_The_Time)
                            if self.pState == 'normal':
                                Analysism.test_Result = 'OK'
                                Analysism.test_Result_status.append(Analysism.test_Result) 
                            
        # 实现场景3：内存一直未超1000MB,但开始与结束的落差值在xxx（divide_The_Value =300mb,后期改为可配置在json文件中）
        elif mNum < int(self.secondaryMaximum) and (int(eNum) - int(sNum)) >= int(self.divide_The_Value) :
            print("实现场景3：内存一直未超1000MB,但开始与结束的实际落差值在%sMB,已超过KPI: %s MB"%((int(eNum) - int(sNum)),self.divide_The_Value))
            self.write_to_time()
            Analysism.sheetNameList.append(self.name)
            if self.pState == 'normal' :
                Analysism.test_Result = 'NG'
                Analysism.test_Result_status.append(Analysism.test_Result)  
            Analysism.error_running_time.append("内存一直未超1000MB,但开始与结束的实际落差值在%sMB,已超过KPI:%sMB"%((int(eNum) - int(sNum)),self.divide_The_Value))
            # 取出存在落差的全部数据创新sheet,并创建柱状图
            if stmpLine <= ftmpLine and ftmpLine > stmpLine:
                with open(self.readPath,'r',encoding='UTF-8',errors="ignore") as read_file:
                    for xline in read_file:
                        xline = xline.strip('\n')
                        dtvList.append(xline)

        # 实现场景4：判断峰值或结束值是未超1024MB，且未长时间1000MB和未超开始结束落差值
        else:
            Analysism.error_running_time.append('无')
            if self.pState == 'normal':
                Analysism.test_Result = 'OK'
                Analysism.test_Result_status.append(Analysism.test_Result) 
            elif self.pState == 'NA':
                Analysism.test_Result = 'NA'
                Analysism.test_Result_status.append(Analysism.test_Result)
            elif self.pState == '第三方NG':
                Analysism.test_Result = '第三方NG'
                Analysism.test_Result_status.append(Analysism.test_Result)

            print("实现场景4：本次内存峰值和结束值不存在超%sMB的测试场景"%(self.KPI))

        return(self.readPath,self.writePath,Analysism.starNumlist,Analysism.endNumlist,Analysism.maxNumlist,Analysism.test_Result_status,Analysism.divide_Time_list)
    # 创建出现问题sheet数量
    def write_to_time(self):
        Analysism.sheetcount +=1
        return(Analysism.sheetNameList)
# 自定义的内存error创建excel内容和数据
def write_to_excel(sheetnamelist,readPath,writePath,timestamp,error_running_time):
    alist = ()             # 放置年月日的元组
    blist = ()            # 放置小时的元组
    flist = ()            # 放置内存结果的元组
    originalname = ''     # 放置log原始名称

    atimestamp = ()
    aerror_running_time = ()

    headings = ['年月日','小时','内存值(MB)']
    print('--write_to_excel-路径-%s'%writePath)

    workbook = xlsxwriter.Workbook(writePath, {'strings_to_numbers':False})
    for sheetindex in range(len(sheetnamelist)):
        index = 0 

        with open(KeyType.testconfigJsonPath,'r',encoding='UTF-8') as c:
            config = json.load(c)
            for d in (config.keys()):
                if d != "Output_Path" and d != "Valgrind_File" and d != "WDX_Output_Path" and d != "MEMkPI" \
                    and d != "SecondaryMaximum" and d != "DivideTheValue" and d != "DivideTheTime" and d != "mailpassCc":
                    data_perison = config.get(d)
                    for item in data_perison.keys():
                        print(item)
                        if item == "grade":
                            data_grade = data_perison["grade"]
                            data_name = data_perison["name"]
                            data_startTime = data_grade['startTime']
                            data_endTime = data_grade['endTime']
                            data_setupFileName = data_grade['setupFilePath']
                            data_setupFilePath = os.path.join(os.path.abspath(os.path.dirname(__file__)),data_setupFileName).replace("\\",'/')
                            originalname = data_setupFileName

                            if data_name == sheetnamelist[sheetindex]:
                                with open(data_setupFilePath,'r',encoding='UTF-8',errors="ignore") as readline:
                                    aalist = []
                                    bblist = []
                                    fflist = []

                                    for kline in readline:
                                        if "BEGIN" in kline:
                                            continue
                                        elif len(kline) >= 84 and KeyType.keyMXNavi in kline:
                                            kline = kline.strip('\n').split()
                                            yearline = (kline[0]).strip()
                                            aalist.append(yearline.strip('['))
                                            hourline = (kline[1]).strip()
                                            bblist.append(hourline.strip(']'))
                                            if kline[5] != None:
                                                a = (kline[5].strip())
                                                fflist.append(int(a[:-1]))
                                            else:
                                                break
                                        elif "END" in kline:
                                            continue
                                        else:
                                            print("数据不满足84位==%s，且不是正确有效数据"%(kline))
                                workbooksheet = workbook.add_worksheet(sheetnamelist[sheetindex])
                                workbooksheet.write_row('A1',headings)
                                # 可变对象转换为不可变对应作为函数的默认值（字典,集合,列表等等对象是不适合作为函数默认值的）
                                alist = tuple(aalist)
                                blist = tuple(bblist)
                                flist = tuple(fflist)
                                workbooksheet.write_column('A2',alist)
                                workbooksheet.write_column('B2',blist)
                                workbooksheet.write_column('C2',flist)
                                workbooksheet.write('D1',originalname )

                                index += 1  
                                #加入数据分析曲线图 
                                categoriesLen = len(blist)
                                valuesLen = len(flist)
    
                                chart_col = workbook.add_chart({'type':'line'})
                                chart_col.add_series(
                                    {
                                    # 内存数据名称和单位
                                    'name':'={sheet_name}!$C$1'.format(sheet_name = sheetnamelist[sheetindex]),
                                    # X轴时间范围
                                    'categories':'= {sheet_name}!$B$2:$B${end}'.format(sheet_name = sheetnamelist[sheetindex],end = categoriesLen),
                                    # Y轴内存曲线
                                    'values':'= {sheet_name}!$C$2:$C${end}'.format(sheet_name = sheetnamelist[sheetindex],end = valuesLen),
                                    'line':{'color':'red'},
                                    }
                                )
                                chart_col.height = 600
                                chart_col.width = 960
                                chart_col.set_title({'name':'稳定性测试'})
                                chart_col.set_x_axis({'name':'运行时间'})
                                chart_col.set_y_axis({'name':'内存值'})
                                chart_col.set_style(1)
                                # 放置位置
                                workbooksheet.insert_chart('E2',chart_col,{'x_offset':25,'y_offset':10})

    workbook.close() 

# 自定义生成汇总Excel表格（已稳定性结果为模板）读已知文档
def readExcel(data_wdx_path,sheet_name,startTimeyear,startTimehour,endTimeyear,endTimehour,dataPath,name_data,data_startNum,data_endNum,\
    data_maxNum,test_Result,effective_running_time,timestamp,error_running_time,divide_Time_list,Space_occupation_Value):
    oldwb = openpyxl.load_workbook(data_wdx_path)
    oldws = oldwb[sheet_name]
    # 添加备注信息（超1000-1024之间的保持耗时）   暂时取消与36列备注信息和平统计
    # for i in range(1,len(divide_Time_list)+1):
    #     datatest = (divide_Time_list[i-1])
    #     oldws.cell(row = i + 4,column = 22).value = datatest[:-1]
    # 添加运行时间
    for i in range(1,len(effective_running_time)+1):
        datatime = effective_running_time[i-1]
        oldws.cell(row = i + 4,column = 23).value = datatime
    # 添加开始日期
    for i in range(1,len(startTimeyear)+1):
        startTime = startTimeyear[i-1]
        oldws.cell(row = i + 4,column = 24).value = startTime
    # 添加开始时间
    for i in range(1,len(startTimehour)+1):
        startTime = startTimehour[i-1]
        oldws.cell(row = i + 4,column = 25).value = startTime
    # 添加开始内存
    for i in range(1,len(data_startNum)+1):
        datastart = data_startNum[i-1]
        oldws.cell(row = i + 4,column = 26).value = datastart
    # 添加结束日期
    for i in range(1,len(endTimeyear)+1):
        startTime = endTimeyear[i-1]
        oldws.cell(row = i + 4,column = 27).value = startTime
    # 添加结束时间
    for i in range(1,len(endTimehour)+1):
        startTime = endTimehour[i-1]
        oldws.cell(row = i + 4,column = 28).value = startTime
    # 添加结束内存
    for i in range(1,len(data_endNum)+1):
        dataend = data_endNum[i-1]
        oldws.cell(row = i + 4,column = 29).value = dataend
    # 添加峰值内存
    for i in range(1,len(data_maxNum)+1):
        datamax = data_maxNum[i-1]
        oldws.cell(row = i + 4,column = 31).value = datamax
    # 添加空间占用峰值
    for i in range(1,len(Space_occupation_Value)+1):
        spcmax = Space_occupation_Value[i-1]
        oldws.cell(row = i + 4,column = 32).value = spcmax
    # 添加log路径
    for i in range(1,len(dataPath)+1):
        logpath = dataPath[i-1]
        oldws.cell(row = i + 4,column = 33).value = logpath
    # 添加测试结果状态
    for i in range(1,len(test_Result)+1):
        datatest = test_Result[i-1]
        oldws.cell(row = i + 4,column = 35).value = datatest
    # 添加备注信息（超1G达到时间 和 第一次超1G的时间戳 + 当在1000-1024之间长时间保持 + 未超1G但开始结束的落差值在300MB以上）  
    for i in range(1,len(error_running_time)+1):
        datatest = (error_running_time[i-1])
        oldws.cell(row = i + 4,column = 37).value = datatest[:-1]
    # 添加测试人员姓名（英文）
    for i in range(1,len(name_data)+1):
        name = name_data[i-1]
        oldws.cell(row = i + 4,column = 38).value = name

    oldwb.save(data_wdx_path)
    print('--write_to_excel-路径-%s'%data_wdx_path)

# 自动邮件管理类（去除图片加入附件）
class EmailManager:
    def __init__(self,mailPass,mailpassCc,mailMsg,mailTitle,filesPath):
        self.mailPass = mailPass
        self.mailpassCc = mailpassCc
        self.mailMsg = mailMsg
        self.mailTitle = mailTitle
        self.filesPath = filesPath

    # def excel_save_img(self,img_suffix="png"):
    #     app = xw.App(visible=True, add_book=False)
    #     # 1. 使用 xlwings 的 读取 path 文件 启动
    #     wb = app.books.open(self.filesPath)
    
    #     # 2. 读取 sheet
    #     sht = wb.sheets['常规版本稳定性测试结果']
    
    #     # 3. 获取 行与列
    #     nrow = sht.api.UsedRange.Rows.count
    #     ncol = sht.api.UsedRange.Columns.count
    #     print(nrow)
    #     print(ncol)
    
    #     # 4. 获取有内容的 range
    #     range_val = sht.range(
    #         (1, 1),  # 获取 第一行 第一列
    #         (nrow, ncol)  # 获取 第 nrow 行 第 ncol 列
    #     )
    #     print(range_val.value)
    
    #     # 5. 复制图片区域
    #     range_val.api.CopyPicture()
    
    #     # 6. 粘贴
    #     sht.api.Paste()
    
    #     pic = sht.pictures[0]  # 当前图片
    #     pic.api.Copy()  # 复制图片
    
    #     img = ImageGrab.grabclipboard()  # 获取剪贴板的图片数据
    #     img.save(img_name + "." + img_suffix)  # 保存图片
    #     pic.delete()  # 删除sheet上的图片
    
    #     wb.close()  # 不保存，直接关闭
    #     app.quit()  # 退出

    def sendEmail(self):
        # 使用的邮箱的SMTP服务器地址
        mail_host = "192.168.2.23"
        # 邮箱的地址
        mail_from = "sunc@meixing.com"
        # 发送到的地址
        mail_pass = self.mailPass
        # 发送到的抄送地址
        mail_passCc = self.mailpassCc
        # 发送的信息
        mail_msg = self.mailMsg
        # 发送的邮件标题
        mail_title = self.mailTitle
        # 图片地址
        # pic_path = self.picPath
        # 附件名称
        files_Path = self.filesPath

        #采用related定义内嵌资源的邮件体
        msg = MIMEMultipart('related') 
        msg['Subject'] = Header(mail_title,'utf-8')  
        msg['From'] = mail_from  
        msg['To'] = mail_pass
        msg['Cc'] = mail_passCc

        msg.attach(MIMEText(mail_msg, 'html', 'utf-8'))
        #发送带有多个Excel附件
        if isinstance(files_Path,list):
            for filepath in files_Path:
                ctype, encoding = mimetypes.guess_type(filepath)
                if ctype is None or encoding is not None: 
                    ctype = "application/octet-stream"
                maintype, subtype = ctype.split("/", 1)
                if maintype in['image','audio']:
                    add_attachment(filepath)
                else:

                    baseName = os.path.basename(filepath) 
                    att = MIMEApplication(open(filepath,'rb').read())
                    att.add_header('Content-Disposition', 'attachment', filename=baseName)
                    msg.attach(att)
                    print(filepath, 'added')
        else:
            ctype, encoding = mimetypes.guess_type(files_Path)
            if ctype is None or encoding is not None: 
                ctype = "application/octet-stream"
            else:
                baseName = os.path.basename(files_Path) 
                att = MIMEApplication(open(files_Path,'rb').read())
                att.add_header('Content-Disposition', 'attachment', filename=baseName)
                msg.attach(att)
                print(files_Path, 'added')

        # 指定图片为当前目录
        # fp = open(pic_path, 'rb')
        # msgImage = MIMEImage(fp.read())
        # fp.close()

        # 定义图片 ID，在 HTML 文本中引用
        # msgImage.add_header('Content-ID', '<image1>')
        # msg.attach(msgImage)

        smtp = smtplib.SMTP()
        # 使用标准的25端口连接SMTP服务器时，使用的是明文传输，发送邮件的整个过程可能会被窃听
        smtp.connect(mail_host,25)
        smtp.sendmail(msg['From'],msg['To'].split(',') + msg['Cc'].split(','),msg.as_string())
        smtp.quit()
        print('邮件发送成功！')
    # excel_save_img(self,img_name='1')      
# 自动读取log中开始和结束时间并写入到testconfig文件
def read_time_wirte_json(loglist):
    tempLienNum = 0
    start_time = ''
    delkeylist = []
    with open(KeyType.configJsonPath,'r',encoding='UTF-8') as c:
        config = json.load(c)

        for timex in (config.keys()):
            if timex != "Output_Path" and timex != "Valgrind_File" and timex != "WDX_Output_Path" and timex != "MEMkPI" \
                and timex != "SecondaryMaximum" and timex != "DivideTheValue" and timex != "DivideTheTime" and timex != "mailpassCc" \
                and timex !="SpaceUsageKPI":
                data_perison = config.get(timex) 

                for item in data_perison.keys():
                    if item == "grade":
                        data_grade = data_perison["grade"]
                        data_name = data_perison["name"]
                        data_setupFileName = data_grade['setupFilePath']
                        print(data_setupFileName)
                        # data_setupFilePath = os.path.join(os.path.abspath(os.path.dirname(__file__)),data_setupFileName).replace("\\",'/')
                        data_setupBasePath = two_abs_join(KeyType.timelinePath,timesfile)
                        print(data_setupBasePath)
                        # data = two_abs_join(data_setupBasePath,data_setupFileName)

                        data_grade['setupFilePath'] = two_abs_join(data_setupBasePath,data_setupFileName).replace('\\','/')
                        data_setupFilePath = data_grade['setupFilePath']
                        # print(data_setupFilePath)
                        # 找到换关键字的第一行
                        if data_setupFileName in loglist:
                            print('1')
                            with open(data_setupFilePath,'r',encoding = "UTF-8",errors = "ignore") as read_file:
                                for line in read_file:
                                    if KeyType.keyMXNavi in line:
                                        start_time = line.split()[0].strip('[') + ' ' + line.split()[1].strip(']')
                                        data_grade['startTime'] = start_time
                                        # print('开始时间：%s'%data_grade['startTime'])
                                        break
                                    tempLienNum = tempLienNum + 1
                            # 找到换关键字的最后一行
                            with open(data_setupFilePath,'r',encoding = "UTF-8",errors = "ignore") as read_file:
                                dq = deque(read_file)
                                while dq :
                                    last_row = dq.pop()
                                    if KeyType.keyMXNavi in last_row:
                                        end_time = last_row.split()[0].strip('[') + ' ' + last_row.split()[1].strip(']')
                                        data_grade['endTime'] = end_time
                                        # print('结束时间：%s'%data_grade['endTime'])
                                        break
                        else:
                            delkeylist.append(timex)
        for i in list(delkeylist):
            if i in list(config.keys()):
                print("删除的key : %s"%(i))
                del config[i]
            else:
                print("不删除的key : %s"%(i))

    with open(KeyType.testconfigJsonPath,'w',encoding="utf-8") as wr :
        json.dump(config,wr,indent=4,sort_keys=False,ensure_ascii=False)
# 拼接路径
def two_abs_join(abs1,abs2):
    abs2 = os.fspath(abs2)

    abs2 = os.path.splitdrive(abs2)[1]

    abs2 = abs2.strip("\\/") or abs2
    return os.fspath(os.path.join(abs1,abs2))
# 自动创建文件夹
def mkdir(path):
    # 去除首位空格
    path = path.strip()
    # 去除尾部 \ 符号
    path=path.rstrip("\\")
    # 判断路径是否存在
    # 存在     True
    # 不存在   False
    isExists=os.path.exists(path)
     # 判断结果
    if not isExists:
        # 如果不存在则创建目录
        # 创建目录操作函数
        os.makedirs(path) 
 
        print(path + ' 创建成功')
        return True
    else:
        # 如果目录存在则不创建，并提示目录已存在
        print(path + ' 目录已存在')
        return False
# 遍历所有文件夹
def checkFile(path):
    a = []
    for i in os.listdir(path):
        path2 = os.path.join(path,i)
        if os.path.isdir(path2):
            checkFile(path2)
        else:
            a.append(i)
    return a

if __name__ == '__main__':

    # 获取当前时间
    daytime = datetime.datetime.now().date()
    boll = is_workday(daytime)  # 输出结果为bool值，True为工作日，False为休息日。
    
    wdx_sheet_name = '常规版本稳定性测试结果'
    namelist = []    # sheet页名称汇总
    readPath = ''    # 读取配置路径
    writePath = ''   # 写入内存曲线文档路径

    name_data = []
    start_time_data_year = []  
    start_time_data_hour = []
    end_time_data_year = []
    end_time_data_hour = []
    data_setupFP = []
    data_startNum = []
    data_endNum = []
    data_maxNum = []

    ok_or_ng = []                              # 测试结果状态收集
    mail_path = []                            # 放置测试人员邮件收集
    placingState = []                        # 测试结果状态收集

    Space_occupation_field = []              # 空间占用收集
    # Space_occupation_Value = ''               # 空间占用最大值
    # Space_occupation_Value_Str = ''           # 空间占用最大值 + 描述
    Space_occupation_Value = []
    effective_running_time = []
    timestamp = ()
    error_running_time = ()
    divide_Time_list = [] 
    Files=[]                             # 测试结果附件名称汇总

    mail_Pass = []                     # 不包含的抄送的全部收件人员list （由全员发送改为只针对存在问题的人员发送）
    mail_Error_Pass = []              # 指定的报错的收件人
    mail_passCc = []                 # 邮件抄送人员list
    mail_Pass_str = ''              # 收件人员list拼接成字符串格式（由全员发送改为只针对存在问题的人员发送）
    mail_Error_Pass_str = ''       # 指定的报错的收件人拼成字符串格式
    mail_passCc_str = ''          # 邮件抄送人员list拼接成字符串格式
    mail_Pass_regulator = 'sunc@meixing.com'     # 邮件收件人管理者
    mailTitle = "【CNS3.0_SOP2_MA】稳定性log分析及填写报告-反馈"
    loglist = []                # 存在log汇总
    # 判断稳定性放置log文件夹内是否为空
    # years = datetime.datetime.now().year
    # months = datetime.datetime.now().month
    # day = datetime.datetime.now().day
    # if len(str(day)) == 1:
    #     day = str(0) + str(day)
    # else:
    #     day = str(day)
    # timesfile  = str(years) + str(months) + day

    # 获取当前时间
    daytime = datetime.datetime.now().date()

    str_daytime = str(daytime)
    timesfile  = str_daytime.replace('-','')

    tname = two_abs_join(KeyType.timelinePath,timesfile)

    if os.path.exists(tname):
        print('%s 文件夹已存在'%(tname))
    else:
        print('%s 文件夹不存在'%(tname))
        mkdir(tname)
        mailMsg = '''
            <p><b>今日文件夹已自动创建完毕，请提醒小伙伴们更新稳定性log！</b></p>
            <p>稳定性结果更新地址：<a href="\\192.168.2.22\cns3.0_sop2_ma\04.C Sample\03.非功能测试\稳定性测试\MX_AnalysisMemoryLog\output">\\\\192.168.2.22\cns3.0_sop2_ma\04.C Sample\03.非功能测试\稳定性测试\MX_AnalysisMemoryLog\output</a></p>
            <p>稳定性结果XshellLog上传路径：<a href="\\192.168.2.22\cns3.0_sop2_ma\04.C Sample\3.非功能测试\稳定性测试">\\\\192.168.2.22\cns3.0_sop2_ma\04.C Sample\03.非功能测试\稳定性测试\</a></p>
            '''

        manager = EmailManager(mail_Pass_regulator,mail_passCc_str,mailMsg,mailTitle,Files)
        manager.sendEmail()

    if os.path.getsize(tname):
        print('%s文件夹是空的'%(tname))
        # 发邮件通知大家更新
    else:
        loglist = checkFile(tname)
        print('%s文件夹不是空的'%(tname))
        print(loglist)
        read_time_wirte_json(loglist)   # 自动读取log时间，写入新json文件并运用

     # 读取json 配置文件路径
    with open(KeyType.testconfigJsonPath,'r',encoding='UTF-8') as c:
        config = json.load(c)
        writeFileName = config['Output_Path']
        writeWdxFielName = config['WDX_Output_Path']
        writeWdxFielPath = os.path.join(os.path.abspath(os.path.dirname(__file__)),writeWdxFielName).replace('\\','/')
        MEMKPI = int(config['MEMkPI'])
        secondaryMaximum = config['SecondaryMaximum']
        divide_The_Value = config['DivideTheValue']
        divide_The_Time = config['DivideTheTime']
        # 收集邮件添附件名称
        Files.append(writeFileName)
        Files.append(writeWdxFielName)
        # 自动收集配置项中抄送人员名称循环加入邮件抄送人员list拼接成字符串格式
        for mailvalues in (config['mailpassCc'].values()):
            mail_passCc.append(mailvalues)
        # 邮件抄送人员list拼接成字符串格式
        # mail_passCc_str = ','.join(mail_passCc)

        #判断 Output_Path 文件夹是否存在
        if os.path.exists(config['Output_Path']):
            print("Output_Path 已存在")
        else:
            #结果False 就创建文件夹 
            os.makedirs(config['Output_Path'])
        for d in (config.keys()):
            if d != "Output_Path" and d != "Valgrind_File" and d != "WDX_Output_Path" and d != "MEMkPI" \
                and d != "SecondaryMaximum" and d != "DivideTheValue" and d != "DivideTheTime" and d != "mailpassCc" \
                and d !="SpaceUsageKPI":
                data_perison = config.get(d)
                for item in data_perison.keys():
                    if item == "grade":
                        data_grade = data_perison["grade"]
                        data_name = data_perison["name"]
                        name_data.append(data_name)
                        mail_path = data_perison["mailPass"]
                        mail_Pass.append(mail_path)

                        data_startTime = data_grade['startTime']
                        data_endTime = data_grade['endTime']

                        start_time_data_year.append(data_startTime.split()[0])
                        start_time_data_hour.append(data_startTime.split()[1])

                        end_time_data_year.append(data_endTime.split()[0])
                        end_time_data_hour.append(data_endTime.split()[1])

                        data_setupFileName = data_grade['setupFilePath']
                        data_setupFilePath = os.path.join(os.path.abspath(os.path.dirname(__file__)),data_setupFileName).replace("\\",'/')
                        data_setupFP.append(data_setupFilePath)
                        placingState = data_grade['placingState']
                        print(data_setupFileName)

                        persion = Analysism(data_name,data_startTime,data_endTime,data_setupFilePath,writeFileName,config['Valgrind_File'],MEMKPI \
                            ,config['SecondaryMaximum'],config['DivideTheValue'],config['DivideTheTime'],placingState,config['SpaceUsageKPI'])
                        persion.check_tmpLine()
                        effective_running_time,timestamp,error_running_time,Space_occupation_field = persion.check_memoryTime()
                        Space_occupation_Value = persion.check_spaceUsage()
                        readPath,writePath,data_startNum,data_endNum,data_maxNum,ok_or_ng,divide_Time_list = persion.check_memorylog()
                        namelist = persion.write_to_time()
                        print('出现问题的最终名单：%s'%namelist)

                    else:
                        print("JSON文件设置项配置错误")  
            else:
                print("不存在Output_Path和Valgrind_File文件夹")

    write_to_excel(namelist,readPath,writePath,timestamp,error_running_time)
    readExcel(writeWdxFielPath,wdx_sheet_name,start_time_data_year,start_time_data_hour,end_time_data_year,end_time_data_hour\
        ,data_setupFP,name_data,data_startNum,data_endNum,data_maxNum,ok_or_ng,effective_running_time,timestamp,error_running_time\
        ,divide_Time_list,Space_occupation_Value)

    # 收件人与抄送人自动判断,判断后并发送结果
    if namelist :
        for d in (config.keys()):
            if d != "Output_Path" and d != "Valgrind_File" and d != "WDX_Output_Path" and d != "MEMkPI" \
                and d != "SecondaryMaximum" and d != "DivideTheValue" and d != "DivideTheTime" and d != "mailpassCc":
                data_perison = config.get(d)
                for item in data_perison.keys():
                    if item == "grade":
                        if data_perison["name"] in namelist:
                            mail_Error_Pass.append(data_perison["mailPass"])
                        else:
                            mail_passCc.append(data_perison["mailPass"])
        # 收件人邮箱名重复处理
        if len(mail_Error_Pass) == len(set(mail_Error_Pass)):
            print('邮件名称不重复')
            mail_Error_Pass_str = ','.join(mail_Error_Pass)
        else:
            print('邮件名称重复')
            mail_Error_Pass_str = ','.join(set(mail_Error_Pass))
        # 抄送人邮箱名重复处理
        if len(mail_passCc) == len(set(mail_passCc)):
            print('邮件名称不重复')
            mail_passCc_str = ','.join(mail_passCc)
        else:
            print('邮件名称重复')
            mail_passCc_str = ','.join(set(mail_passCc))
        # mail_Error_Pass_str = ','.join(mail_Error_Pass)
        print('出现问题时的收件人地址：%s'%(mail_Error_Pass_str))
        print('其他OK的收件人地址和抄送地址：%s'%(mail_passCc_str))

        if boll :
            mailMsg = '''
            <p><b>当天的稳定性log分析及填写报告已生成，请参看附件！</b></p>
            <p>稳定性测试结果存在<b><font color="red">NG</font></b>，以上收件人请注意！</p>
            <p>稳定性结果更新地址：<a href="\\192.168.2.22\cns3.0_sop2_ma\04.C Sample\03.非功能测试\稳定性测试\MX_AnalysisMemoryLog\output">\\\\192.168.2.22\cns3.0_sop2_ma\04.C Sample\03.非功能测试\稳定性测试\MX_AnalysisMemoryLog\output</a></p>
            <p>稳定性结果XshellLog上传路径：<a href="\\192.168.2.22\cns3.0_sop2_ma\04.C Sample\03.非功能测试\稳定性测试">\\\\192.168.2.22\cns3.0_sop2_ma\04.C Sample\03.非功能测试\稳定性测试\</a></p>
            '''
            manager = EmailManager(mail_Error_Pass_str,mail_passCc_str,mailMsg,mailTitle,Files)
            manager.sendEmail()
        else :
            print('今天 %s 是节假日,无需发邮件'%daytime)
    else:
        # 抄送人邮箱如与管理者重复则移除操作
        mail_Pass.remove(mail_Pass_regulator)
        mail_passCc_str = ','.join(mail_passCc)
        # 抄送人邮箱名重复处理
        if len(mail_Pass) == len(set(mail_Pass)):
            print('邮件不重复')
            mail_Pass_str = ','.join(mail_Pass)
        else:
            print('邮件重复')
            mail_Pass_str = ','.join(set(mail_Pass))

        mail_full_pass = mail_Pass_str + ',' + mail_passCc_str
        print('全部无问题时发送管理者收件人：%s'%(mail_Pass_regulator))
        print('配置项中原有抄送地址+除管理者外其他收件人：%s'%(mail_full_pass))
        if boll :
            mailMsg = '''
            <p><b>当天的稳定性log分析及填写报告已生成，请参看附件！</b></p>
            <p>稳定性测试结果全部<b>OK</b></p>
            <p>稳定性结果更新地址：<a href="\\192.168.2.22\cns3.0_sop2_ma\04.C Sample\03.非功能测试\稳定性测试\MX_AnalysisMemoryLog\output">\\\\192.168.2.22\cns3.0_sop2_ma\04.C Sample\03.非功能测试\稳定性测试\MX_AnalysisMemoryLog\output</a></p>
            <p>稳定性结果XshellLog上传路径：<a href="\\192.168.2.22\cns3.0_sop2_ma\04.C Sample\03.非功能测试\稳定性测试">\\\\192.168.2.22\cns3.0_sop2_ma\04.C Sample\03.非功能测试\稳定性测试\</a></p>
            '''
            okFiles = Files[1]
            manager = EmailManager(mail_Pass_regulator,mail_full_pass,mailMsg,mailTitle,okFiles)
            manager.sendEmail()
        else :
            print('今天 %s 是节假日,无需发邮件'%daytime)


    







